import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import openpyxl

json_input_name = str(input("Enter the name of the JSON file: "))

# Load Trivy JSON report
with open(json_input_name, "r", encoding="utf-8") as file:
    data = json.load(file)

# Extract general metadata
artifact_info = {
    "Artifact Name": data.get("ArtifactName", ""),
    "Artifact Type": data.get("ArtifactType", ""),
    "OS Family": data.get("Metadata", {}).get("OS", {}).get("Family", ""),
    "OS Version": data.get("Metadata", {}).get("OS", {}).get("Name", ""),
    "Image ID": data.get("Metadata", {}).get("ImageID", ""),
    "Created At": data.get("CreatedAt", ""),
    "Repo Tags": ", ".join(data.get("Metadata", {}).get("RepoTags", [])),
    "Repo Digests": ", ".join(data.get("Metadata", {}).get("RepoDigests", [])),
    "Image Created": data.get("Metadata", {}).get("ImageConfig", {}).get("created", "")
}

# Extract vulnerabilities
detailed_data = []
if "Results" in data:
    for result in data["Results"]:
        if "Vulnerabilities" in result:
            for vuln in result["Vulnerabilities"]:
                detailed_data.append({
                    "Vulnerability ID": vuln.get("VulnerabilityID", ""),
                    "Package": vuln.get("PkgName", ""),
                    "Installed Version": vuln.get("InstalledVersion", ""),
                    "Fixed Version": vuln.get("FixedVersion", ""),
                    "Status": vuln.get("Status", ""),
                    "Severity Level": vuln.get("Severity", ""),
                    "Severity Source": vuln.get("SeveritySource", ""),
                    "Primary URL": vuln.get("PrimaryURL", ""),
                    "Title": vuln.get("Title", ""),
                    "Description": vuln.get("Description", ""),
                    "References": ", ".join(vuln.get("References", [])),
                    "Published Date": vuln.get("PublishedDate", "")
                })

# Extract history data
history_data = []
history_entries = data.get("Metadata", {}).get("ImageConfig", {}).get("history", [])
for entry in history_entries:
    history_data.append({
        "Created At": entry.get("created", ""),
        "Is Empty Layer?": entry.get("empty_layer", False),
        "Created By": entry.get("created_by", "")
    })

# Convert to DataFrames
summary_df = pd.DataFrame([artifact_info])
vulnerabilities_df = pd.DataFrame(detailed_data)
history_df = pd.DataFrame(history_data)

# Create Excel workbook
wb = Workbook()
ws_summary = wb.active
ws_summary.title = "Summary"
ws_history = wb.create_sheet(title="History")
ws_vulns = wb.create_sheet(title="Vulnerabilities")

# Populate Summary Sheet
for row_idx, (key, value) in enumerate(artifact_info.items(), start=1):
    ws_summary.cell(row=row_idx, column=1, value=key).font = Font(bold=True)
    ws_summary.cell(row=row_idx, column=2, value=value)
    ws_summary.cell(row=row_idx, column=1).fill = PatternFill(start_color="FFC000", fill_type="solid")

# Autofit columns in Summary Sheet
for col in ws_summary.columns:
    max_length = 0
    column = col[0].column_letter  # Get column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws_summary.column_dimensions[column].width = adjusted_width

# Populate History Sheet
for r_idx, row in enumerate(dataframe_to_rows(history_df, index=False, header=True), start=1):
    for c_idx, value in enumerate(row, start=1):
        cell = ws_history.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFA07A", fill_type="solid")
# Apply autofilter to the "History" sheet (just showing the filter icon)

ws_history.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(history_df.columns))}1"

# Autofit columns in History Sheet
history_cols_to_autofit = ["Created At", "Created By", "Is Empty Layer?"]
for col in history_cols_to_autofit:
    col_idx = history_df.columns.get_loc(col) + 1
    max_length = 0
    for row in ws_history.iter_rows(min_col=col_idx, max_col=col_idx):
        for cell in row:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
    adjusted_width = (max_length + 2)
    ws_history.column_dimensions[ws_history.cell(1, col_idx).column_letter].width = adjusted_width

# Populate Vulnerabilities Sheet
for r_idx, row in enumerate(dataframe_to_rows(vulnerabilities_df, index=False, header=True), start=1):
    for c_idx, value in enumerate(row, start=1):
        cell = ws_vulns.cell(row=r_idx, column=c_idx, value=value)
        if r_idx == 1:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFC000", fill_type="solid")

        # Apply custom styling based on Severity Level
        if r_idx > 1:  # Skip header row
            severity = vulnerabilities_df.iloc[r_idx - 2][
                "Severity Level"]  # r_idx-2 because dataframe index is zero-based
            if severity == "CRITICAL":
                cell.fill = PatternFill(start_color="FF0000", fill_type="solid")  # Bright Red for Critical
                cell.font = Font(color="FFFFFF")  # White font for contrast
            elif severity == "HIGH":
                cell.fill = PatternFill(start_color="FFCCCC", fill_type="solid")  # Light Red for High
                cell.font = Font(color="000000")  # Black font for better visibility

# Apply autofilter to the "Vulnerabilities" sheet (just showing the filter icon)
ws_vulns.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(vulnerabilities_df.columns))}1"

# Autofit selected columns in Vulnerabilities Sheet
vuln_cols_to_autofit = [
    "Vulnerability ID", "Published Date", "Severity Source",
    "Status", "Installed Version", "Fixed Version", "Severity Level"
]
for col in vuln_cols_to_autofit:
    col_idx = vulnerabilities_df.columns.get_loc(col) + 1  # get index of column + 1 for openpyxl
    max_length = 0
    for row in ws_vulns.iter_rows(min_col=col_idx, max_col=col_idx):
        for cell in row:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
    adjusted_width = (max_length + 2)
    ws_vulns.column_dimensions[ws_vulns.cell(1, col_idx).column_letter].width = adjusted_width

from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# Post-processing for Unique Vulnerabilities
unique_vulnerabilities_df = vulnerabilities_df[
    ["Package", "Installed Version", "Fixed Version", "Status", "Severity Level"]].drop_duplicates()

# Create a new sheet for Unique Vulnerabilities
ws_unique_vulns = wb.create_sheet(title="Unique Vulnerabilities")

# Populate the new sheet with the unique data
ws_unique_vulns.append(["Package", "Installed Version", "Fixed Version", "Status", "Severity Level"])

# Apply proper Excel formatting for headers
for col_idx, col_name in enumerate(["Package", "Installed Version", "Fixed Version", "Status", "Severity Level"],
                                   start=1):
    cell = ws_unique_vulns.cell(row=1, column=col_idx, value=col_name)
    cell.font = Font(bold=True)
    cell.fill = PatternFill(start_color="FFFF00", fill_type="solid")  # Yellow background for headers

# Add the data rows
for r_idx, row in enumerate(dataframe_to_rows(unique_vulnerabilities_df, index=False, header=False), start=2):
    for c_idx, value in enumerate(row, start=1):
        cell = ws_unique_vulns.cell(row=r_idx, column=c_idx, value=value)

    # Apply custom styling based on Severity Level for the entire row
    severity = unique_vulnerabilities_df.iloc[r_idx - 2]["Severity Level"]  # Get the severity level from the dataframe
    if severity == "CRITICAL":
        fill = PatternFill(start_color="FF0000", fill_type="solid")  # Bright Red for Critical
        font = Font(color="FFFFFF")  # White font for contrast
    elif severity == "HIGH":
        fill = PatternFill(start_color="FFCCCC", fill_type="solid")  # Light Red for High
        font = Font(color="000000")  # Black font for better visibility
    else:
        fill = PatternFill(start_color="FFFFFF", fill_type="solid")  # No color for undefined severity
        font = Font(color="000000")  # Black font

    # Apply the styles to the entire row
    for c_idx in range(1, 6):  # Loop over all columns for the current row
        cell = ws_unique_vulns.cell(row=r_idx, column=c_idx)
        cell.fill = fill
        cell.font = font

# Apply autofilter to the "Unique Vulnerabilities" sheet (just showing the filter icon)
ws_unique_vulns.auto_filter.ref = f"A1:{get_column_letter(len(unique_vulnerabilities_df.columns))}1"

# Autofit columns in Unique Vulnerabilities Sheet
for col in ws_unique_vulns.columns:
    max_length = 0
    column = col[0].column_letter  # Get column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    ws_unique_vulns.column_dimensions[column].width = adjusted_width

ws_stats = wb.create_sheet(title="Unique-Vulnerability-Statistics")

# Populate Overall-Statistics Sheet
total_vulns = len(unique_vulnerabilities_df)
severities_count = unique_vulnerabilities_df["Severity Level"].value_counts().to_dict()

ws_stats.append(["Total Vulnerabilities", total_vulns])
ws_stats.append(["", ""])  # Empty Row
for severity, count in severities_count.items():
    ws_stats.append([severity, count])
    ws_stats.cell(row=ws_stats.max_row, column=1).font = Font(bold=True)
    ws_stats.cell(row=ws_stats.max_row, column=1).fill = PatternFill(start_color="00B0F0", fill_type="solid")

# Autofit columns in Overall-Statistics Sheet
for col in ws_stats.columns:
    max_length = 0
    column = col[0].column_letter  # Get column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 10)
    ws_stats.column_dimensions[column].width = adjusted_width

ws_stats = wb.create_sheet(title="Overall-Statistics")

# Populate Overall-Statistics Sheet
total_vulns = len(vulnerabilities_df)
severities_count = vulnerabilities_df["Severity Level"].value_counts().to_dict()

ws_stats.append(["Total Vulnerabilities", total_vulns])
ws_stats.append(["", ""])  # Empty Row
for severity, count in severities_count.items():
    ws_stats.append([severity, count])
    ws_stats.cell(row=ws_stats.max_row, column=1).font = Font(bold=True)
    ws_stats.cell(row=ws_stats.max_row, column=1).fill = PatternFill(start_color="00B0F0", fill_type="solid")

# Autofit columns in Overall-Statistics Sheet
for col in ws_stats.columns:
    max_length = 0
    column = col[0].column_letter  # Get column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 10)
    ws_stats.column_dimensions[column].width = adjusted_width

# Save the workbook with the new sheet

# remove .json extension from the input file
json_input_name = json_input_name.replace(".json", "")
wb.save(f"{json_input_name}.xlsx")
print("Excel report generated successfully with Unique Vulnerabilities sheet: trivy-report.xlsx")
